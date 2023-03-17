import csv
import os

import openai
import openpyxl

cwd = os.getcwd()
PATH = f"{cwd}\Welltory Test_Python Developer_App Reviews.xlsx"
OPENAI_TOKEN = 'API KEY'


def find_max_row_with_values(sheet):
    max_row = 0
    for row in sheet.iter_rows(values_only=True):
        if any(row):
            max_row += 1
    return max_row


def read_excel_file(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook["Data"]
    data = []
    column_names = ["email", "review_text", "date", "rate"]
    max_row = find_max_row_with_values(sheet)
    for i in range(2, max_row + 1):
        row_dict = {}
        for j, column_name in enumerate(column_names):
            cell_value = sheet.cell(row=i, column=j + 1).value
            row_dict[column_name] = cell_value
        row_dict['id'] = i - 1
        data.append(row_dict)
    workbook.close()
    return data


def generate_prompt(texts):
    formatted_texts = (
        f"{chr(10).join([f'{i}. {review}' for i, review in enumerate(review_texts, start=1)])}\n"
    )
    prompt = (
        f"Rank the following {len(texts)} reviews from most positive to most negative:\n"
        f"{formatted_texts}"
        f"Return a list of review's numbers, sorted by their ranks "
        f"(positive first)."
    )
    return prompt


def generate_rankings(prompt, token):
    openai.api_key = token
    response = openai.Completion.create(
        engine='text-davinci-002',
        prompt=prompt,
        max_tokens=1024,
        n=1,
        stop=None,
        temperature=0.8,
    )
    res = response['choices'][0]['text'].strip('\n').replace(',', '')
    ranks = [int(r) for r in res.split()]
    return ranks


def sort_data_by_rankings(data, ranks):
    return sorted(data, key=lambda x: ranks.index(x['id']))


def set_ratings_and_format_data(sorted_result_data):
    sorted_result_data = list(reversed(sorted_result_data))
    for d in sorted_result_data:
        del d['id']
    for i, d in enumerate(sorted_result_data):
        d['rate'] = i + 1
    for d in sorted_result_data:
        d['date'] = d['date'].strftime("%d.%m.%Y")


def set_rates_to_excel_file(path, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["Data"]
    start_row = 2

    for i in range(start_row, len(data) + start_row):
        rate_cell = sheet.cell(row=i, column=4)
        row_text = sheet.cell(row=i, column=2).value
        row_email = sheet.cell(row=i, column=1).value

        corresponding_rate = [
            d['rate'] for d in data if
            d['review_text'] == row_text and d['email'] == row_email
        ][0]
        rate_cell.value = corresponding_rate

    workbook.save(path)
    workbook.close()


def write_to_csv(data, filename):
    with open(filename, 'w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=data[0].keys())
        writer.writeheader()
        for d in data:
            writer.writerow(d)


if __name__ == '__main__':
    result_data = read_excel_file(PATH)
    review_texts = [row["review_text"] for row in result_data]

    rankings = generate_rankings(
        prompt=generate_prompt(review_texts),
        token=OPENAI_TOKEN
    )

    data_sorted = sort_data_by_rankings(result_data, rankings)

    set_ratings_and_format_data(data_sorted)

    set_rates_to_excel_file(PATH, data_sorted)

    write_to_csv(data_sorted, 'reviews_analyzed.csv')
